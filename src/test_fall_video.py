"""
=============================================================================
  COS30018 — Automated Fall Detection
  Comparative Video Testing Script
  Supports: One-Step | Two-Step | Temporal Smoothing
=============================================================================

MODES:
  1-step baseline (matches your existing results at 0.35 threshold):
    python src/test_video.py

  1-step WITH temporal smoothing:
    python src/test_video.py --temporal

  2-step (person detector + fall classifier):
    python src/test_video.py --two-step

  2-step WITH temporal smoothing:
    python src/test_video.py --two-step --temporal

  Run ALL 4 modes and produce comparison report:
    python src/test_video.py --compare-all

OUTPUT per run — results_video/<video_name>/<mode>/
  annotated_<name>.mp4        full annotated video
  fall_clips/                 one clip per distinct fall event
  detections.xlsx             per-frame detection log (includes consecutive col)
  timeline.png / confidence_dist.png / class_breakdown.png / summary_dashboard.png

results_video/comparison_report.txt   side-by-side metrics across all 4 modes
=============================================================================
"""

import sys, argparse, warnings
from pathlib import Path
from datetime import datetime, timedelta
from collections import deque
import time as _time

warnings.filterwarnings("ignore")

MISSING = []
for pkg, imp in [
    ("opencv-python","cv2"),("numpy","numpy"),("pandas","pandas"),
    ("matplotlib","matplotlib"),("seaborn","seaborn"),
    ("openpyxl","openpyxl"),("ultralytics","ultralytics"),
]:
    try: __import__(imp)
    except ImportError: MISSING.append(pkg)

if MISSING:
    print(f"\n[ERROR] pip install {' '.join(MISSING)}\n"); sys.exit(1)

import cv2, numpy as np, pandas as pd
import matplotlib; matplotlib.use("Agg")
import matplotlib.pyplot as plt, matplotlib.gridspec as gridspec
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from ultralytics import YOLO

# ── Config ────────────────────────────────────────────────────────────────────
FALL_MODEL_PATH   = "src/models/techcare.pt"
PERSON_MODEL_PATH = "src/models/retrained.pt"
VIDEOS_DIR        = "src/videos/videos"
RESULTS_DIR       = "src/results_video"

CONF_THRESHOLD    = 0.35   # baseline — matches your existing results
IMG_SIZE          = 640
DEVICE            = 0
FRAME_SKIP        = 3

CONFIRMED_THRESH  = 0.70
POSSIBLE_THRESH   = 0.35
FALL_PERSIST      = 5      # frames required
POSSIBLE_PERSIST  = 3
ALERT_HOLD_SEC    = 3.0

PERSON_CONF       = 0.45
PERSON_CLS        = "person"

WRITE_VIDEO       = True
OUTPUT_CODEC      = "mp4v"
WRITE_CLIPS       = True
CLIP_PAD_SEC      = 2.0
MAX_CLIPS         = 20
SAVE_FRAMES       = True
MAX_FRAMES        = 30
EVENT_GAP         = 3.0
ROLL_WIN_SEC      = 10
FLASH_N           = 8

CLS_FALL="Fall"; CLS_SIT="Sit"; CLS_WALK="Walk"
BOX_COLS = {CLS_FALL:(0,40,200), CLS_SIT:(0,165,255), CLS_WALK:(60,200,80)}
HEX_COLS = {CLS_FALL:"#C0392B",  CLS_SIT:"#E67E22",   CLS_WALK:"#2E7D52"}

C_RED="#C0392B"; C_GREEN="#2E7D52"; C_BLUE="#2E5496"
C_AMBER="#D4821A"; C_GRAY="#6B7280"; C_LIGHT="#F0F4F8"; C_DARK="#1A1A2E"

VIDEO_EXT = {".mp4",".avi",".mov",".mkv",".webm",".m4v"}

# ── Helpers ───────────────────────────────────────────────────────────────────
def banner(t): print(f"\n{'─'*62}\n  {t}\n{'─'*62}")
def fmt(s):
    td=timedelta(seconds=int(s)); tot=int(td.total_seconds())
    h,r=divmod(tot,3600); m,sc=divmod(r,60)
    return f"{h:02d}:{m:02d}:{sc:02d}" if h else f"{m:02d}:{sc:02d}"

def get_videos(d):
    p=Path(d)
    if not p.exists(): print(f"[ERROR] {d}"); sys.exit(1)
    vs=sorted([f for f in p.iterdir() if f.suffix.lower() in VIDEO_EXT])
    if not vs: print(f"[ERROR] no videos in {d}"); sys.exit(1)
    return vs

def check_fall(boxes):
    if not boxes: return False,"none",0.0
    bc=0.0; bcls="none"; fc=0.0; isf=False
    for b in boxes:
        if b["conf"]>bc: bc=b["conf"]; bcls=b["class_name"]
        if b["class_name"]==CLS_FALL:
            isf=True
            if b["conf"]>fc: fc=b["conf"]
    return isf, bcls, round(fc if isf else bc, 4)

# ── Temporal Smoother ─────────────────────────────────────────────────────────
class TemporalSmoother:
    def __init__(self):
        self.consec=0; self.hist=deque(maxlen=30); self.last_t=-999.0
    def reset(self):
        self.consec=0; self.hist.clear(); self.last_t=-999.0
    def update(self, is_fall, conf):
        if is_fall: self.consec+=1; self.hist.append(conf)
        else:       self.consec=0
        if is_fall and self.consec>=FALL_PERSIST:
            if conf>=CONFIRMED_THRESH:   lvl="confirmed"; self.last_t=_time.time()
            elif conf>=POSSIBLE_THRESH:  lvl="possible";  self.last_t=_time.time()
            else:                        lvl="none"
        elif is_fall and self.consec>=POSSIBLE_PERSIST:
            lvl="possible" if conf>=POSSIBLE_THRESH else "none"
            if lvl=="possible": self.last_t=_time.time()
        else: lvl="none"
        if lvl=="none" and (_time.time()-self.last_t)<ALERT_HOLD_SEC: lvl="possible"
        return lvl, conf if is_fall else 0.0

# ── Two-Step Detector ─────────────────────────────────────────────────────────
class TwoStep:
    def __init__(self, pm, fm):
        self.pm=pm; self.fm=fm
    def predict(self, frame):
        s1=self.pm.predict(source=frame,conf=PERSON_CONF,imgsz=IMG_SIZE,device=DEVICE,verbose=False)[0]
        pboxes=[]
        if s1.boxes is not None:
            for box in s1.boxes:
                if self.pm.names[int(box.cls[0])].lower()==PERSON_CLS:
                    pboxes.append({"conf":float(box.conf[0]),"xyxy":box.xyxy[0].cpu().numpy().tolist()})
        if not pboxes: return []
        h,w=frame.shape[:2]; out=[]
        for pb in pboxes:
            x1,y1,x2,y2=[int(v) for v in pb["xyxy"]]
            x1=max(0,x1);y1=max(0,y1);x2=min(w,x2);y2=min(h,y2)
            if (x2-x1)<10 or (y2-y1)<10: continue
            crop=frame[y1:y2,x1:x2]
            s2=self.fm.predict(source=crop,conf=CONF_THRESHOLD,imgsz=IMG_SIZE,device=DEVICE,verbose=False)[0]
            if s2.boxes is not None and len(s2.boxes)>0:
                confs=s2.boxes.conf.cpu().numpy(); clss=s2.boxes.cls.cpu().numpy().astype(int)
                best=confs.argmax()
                out.append({"class_name":self.fm.names[clss[best]],"conf":float(confs[best]),"xyxy":pb["xyxy"]})
            else:
                out.append({"class_name":"Walk","conf":pb["conf"],"xyxy":pb["xyxy"]})
        return out

# ── Drawing ───────────────────────────────────────────────────────────────────
def draw_boxes(frame, boxes, is_fall, mode_lbl):
    out=frame.copy()
    for b in boxes:
        cls=b["class_name"]; conf=b["conf"]
        x1,y1,x2,y2=map(int,b["xyxy"])
        color=BOX_COLS.get(cls,(200,200,200))
        thick=3 if cls==CLS_FALL else 2
        cv2.rectangle(out,(x1,y1),(x2,y2),color,thick)
        lbl=f"{cls} {conf:.0%}"
        (lw,lh),_=cv2.getTextSize(lbl,cv2.FONT_HERSHEY_SIMPLEX,0.55,1)
        cv2.rectangle(out,(x1,y1-lh-8),(x1+lw+6,y1),color,-1)
        cv2.putText(out,lbl,(x1+3,y1-4),cv2.FONT_HERSHEY_SIMPLEX,0.55,(255,255,255),1,cv2.LINE_AA)
        if cls==CLS_FALL:
            tick=20
            for cx,cy,dx,dy in [(x1,y1,1,1),(x2,y1,-1,1),(x1,y2,1,-1),(x2,y2,-1,-1)]:
                cv2.line(out,(cx,cy),(cx+dx*tick,cy),color,3)
                cv2.line(out,(cx,cy),(cx,cy+dy*tick),color,3)
    if is_fall:
        badge="  FALL DETECTED  "
        (bw,bh),_=cv2.getTextSize(badge,cv2.FONT_HERSHEY_DUPLEX,0.70,1)
        bx=(out.shape[1]-bw)//2
        cv2.rectangle(out,(bx-4,8),(bx+bw+4,bh+24),BOX_COLS[CLS_FALL],-1)
        cv2.putText(out,badge,(bx,bh+16),cv2.FONT_HERSHEY_DUPLEX,0.70,(255,255,255),1,cv2.LINE_AA)
    cv2.putText(out,f"COS30018  |  {mode_lbl}",(8,20),cv2.FONT_HERSHEY_SIMPLEX,0.42,(200,200,200),1,cv2.LINE_AA)
    return out

def draw_hud(frame, ts, dur, is_fall, dom, conf, events, flash):
    out=frame.copy(); h,w=out.shape[:2]
    if flash>0:
        alpha=0.30*(flash/FLASH_N); ov=out.copy()
        cv2.rectangle(ov,(0,0),(w,h),(0,30,160),-1)
        cv2.addWeighted(ov,alpha,out,1-alpha,0,out)
    sh=52; cv2.rectangle(out,(0,h-sh),(w,h),(20,20,20),-1)
    sm=cv2.FONT_HERSHEY_SIMPLEX; dp=cv2.FONT_HERSHEY_DUPLEX
    cv2.putText(out,f"{fmt(ts)} / {fmt(dur)}",(10,h-sh+20),sm,0.50,(230,230,230),1,cv2.LINE_AA)
    bx,by,bw2,bh2=10,h-sh+30,160,8
    cv2.rectangle(out,(bx,by),(bx+bw2,by+bh2),(60,60,60),-1)
    cv2.rectangle(out,(bx,by),(bx+int(bw2*min(ts/max(dur,1),1.0)),by+bh2),(140,140,140),-1)
    stxt=f"FALL DETECTED ({conf:.0%})" if is_fall else f"MONITORING — {dom.upper()}"
    scol=(0,40,200) if is_fall else (80,180,60)
    (tw,_),_=cv2.getTextSize(stxt,dp,0.52,1)
    cv2.putText(out,stxt,((w-tw)//2,h-sh+22),dp,0.52,scol,1,cv2.LINE_AA)
    b2x=w-190; b2y=h-sh+14; b2w=140; b2h=10
    cv2.putText(out,f"CONF: {conf:.0%}",(b2x,b2y+8),sm,0.45,(230,230,230),1,cv2.LINE_AA)
    cv2.rectangle(out,(b2x,b2y+14),(b2x+b2w,b2y+14+b2h),(60,60,60),-1)
    cfill=min(int(b2w*conf),b2w)
    ccol=(0,180,60) if conf>=0.70 else ((0,165,255) if conf>=0.50 else (0,80,200))
    cv2.rectangle(out,(b2x,b2y+14),(b2x+cfill,b2y+14+b2h),ccol,-1)
    (ew,_),_=cv2.getTextSize(f"FALLS: {events}",sm,0.50,1)
    cv2.putText(out,f"FALLS: {events}",(w-ew-8,h-sh+46),sm,0.50,(230,230,230),1,cv2.LINE_AA)
    return out

# ── Event segmentation + clips ────────────────────────────────────────────────
def segment_events(df):
    if df.empty: return []
    fdf=df[df["is_fall"]==1].copy()
    if fdf.empty: return []
    evs=[]; eid=0; rows=fdf.sort_values("timestamp_s").to_dict("records"); i=0
    while i<len(rows):
        s=rows[i]["timestamp_s"]; e=s; pk=rows[i]["confidence"]; c=1
        while i+1<len(rows):
            if rows[i+1]["timestamp_s"]-rows[i]["timestamp_s"]>EVENT_GAP: break
            i+=1; e=rows[i]["timestamp_s"]; pk=max(pk,rows[i]["confidence"]); c+=1
        eid+=1; evs.append({"event_id":eid,"start_s":s,"end_s":e,"peak_conf":round(pk,4),"frame_count":c}); i+=1
    return evs

def write_clips(vpath, evs, clips_dir, fps, fw, fh, model, dur, two_step_det, mode_lbl):
    if not evs: return
    clips_dir.mkdir(exist_ok=True)
    codec=cv2.VideoWriter_fourcc(*OUTPUT_CODEC); n=0
    for ev in evs:
        if n>=MAX_CLIPS: break
        cs=max(0.0,ev["start_s"]-CLIP_PAD_SEC); ce=min(dur,ev["end_s"]+CLIP_PAD_SEC)
        sf=int(cs*fps); ef=int(ce*fps)
        clip=clips_dir/f"fall_{ev['event_id']:03d}_{fmt(ev['start_s']).replace(':','-')}.mp4"
        cap=cv2.VideoCapture(str(vpath)); cap.set(cv2.CAP_PROP_POS_FRAMES,sf)
        wr=cv2.VideoWriter(str(clip),codec,fps,(fw,fh))
        fi=sf; fl=0
        print(f"  Clip {ev['event_id']:02d} @ {fmt(ev['start_s'])}",end="",flush=True)
        while fi<=ef:
            ret,frame=cap.read()
            if not ret: break
            ts=fi/fps
            try:
                if two_step_det: boxes=two_step_det.predict(frame)
                else:
                    r=model.predict(source=frame,conf=CONF_THRESHOLD,imgsz=IMG_SIZE,device=DEVICE,verbose=False)[0]
                    boxes=[{"class_name":model.names[int(b.cls[0])],"conf":float(b.conf[0]),"xyxy":b.xyxy[0].cpu().numpy().tolist()} for b in (r.boxes or [])]
                isf,dom,conf=check_fall(boxes)
                if isf: fl=FLASH_N
                ann=draw_boxes(frame,boxes,isf,mode_lbl)
                out=draw_hud(ann,ts,dur,isf,dom,conf,ev["event_id"],fl)
                if fl>0: fl-=1
                wr.write(out)
            except: wr.write(frame)
            fi+=1
        cap.release(); wr.release(); n+=1; print("  ✓")
    print(f"  {n} clip(s) → {clips_dir.name}/")

# ── Main inference pass ───────────────────────────────────────────────────────
def process_video(fall_model, video_path, out_dir, two_step_det, temporal_on, mode_lbl):
    cap=cv2.VideoCapture(str(video_path))
    if not cap.isOpened(): print(f"  [SKIP] {video_path.name}"); return pd.DataFrame()
    fps=cap.get(cv2.CAP_PROP_FPS) or 30.0
    total=int(cap.get(cv2.CAP_PROP_FRAME_COUNT)); dur=total/fps
    fw=int(cap.get(cv2.CAP_PROP_FRAME_WIDTH)); fh=int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    print(f"\n  {video_path.name}  [{mode_lbl}]  {fw}x{fh}  {fps:.1f}fps  {fmt(dur)}")

    wr=None; apath=None
    if WRITE_VIDEO:
        codec=cv2.VideoWriter_fourcc(*OUTPUT_CODEC)
        apath=out_dir/f"annotated_{video_path.stem}.mp4"
        wr=cv2.VideoWriter(str(apath),codec,fps,(fw,fh))
    fdir=out_dir/"annotated_frames"
    if SAVE_FRAMES: fdir.mkdir(exist_ok=True)

    smoother=TemporalSmoother(); rows=[]; fi=0; saved=0
    events=0; flash=0; last_fall_ts=-999.0; boxes=[]; is_fall=False; dom="none"; conf=0.0

    while True:
        ret,frame=cap.read()
        if not ret: break
        fi+=1; ts=fi/fps
        alert_lvl="none"

        if fi%FRAME_SKIP==0:
            try:
                if two_step_det: boxes=two_step_det.predict(frame)
                else:
                    r=fall_model.predict(source=frame,conf=CONF_THRESHOLD,imgsz=IMG_SIZE,device=DEVICE,verbose=False)[0]
                    boxes=[{"class_name":fall_model.names[int(b.cls[0])],"conf":float(b.conf[0]),"xyxy":b.xyxy[0].cpu().numpy().tolist()} for b in (r.boxes or [])]
            except: boxes=[]

            classes={b["class_name"] for b in boxes}
            isf_raw,dom,conf=check_fall(boxes)

            if temporal_on:
                alert_lvl,_=smoother.update(isf_raw,conf)
                is_fall=alert_lvl in ("confirmed","possible")
            else:
                is_fall=isf_raw
                alert_lvl="confirmed" if is_fall else "none"

            if is_fall and (ts-last_fall_ts)>EVENT_GAP:
                events+=1; flash=FLASH_N; last_fall_ts=ts
            elif is_fall: last_fall_ts=ts
            if flash>0: flash-=1

            rows.append({
                "frame_number":fi,"timestamp_s":round(ts,3),"timestamp_str":fmt(ts),
                "is_fall":int(is_fall),"alert_level":alert_lvl,
                "dominant_class":dom,"confidence":conf,
                "consecutive":smoother.consec if temporal_on else (1 if is_fall else 0),
                "classes_seen":", ".join(sorted(classes)) or "none","num_boxes":len(boxes),
                "has_fall":int(CLS_FALL in classes),"has_sit":int(CLS_SIT in classes),"has_walk":int(CLS_WALK in classes),
            })

            if SAVE_FRAMES and is_fall and saved<MAX_FRAMES:
                ann=draw_boxes(frame,boxes,is_fall,mode_lbl)
                cv2.imwrite(str(fdir/f"fall_{fmt(ts).replace(':','-')}_f{fi:06d}.jpg"),ann)
                saved+=1

        if wr:
            ann=draw_boxes(frame,boxes,is_fall,mode_lbl)
            wr.write(draw_hud(ann,ts,dur,is_fall,dom,conf,events,flash))

        pct=min(fi/max(total,1),1.0)
        bar="█"*int(32*pct)+"░"*(32-int(32*pct))
        print(f"\r  [{bar}] {fmt(ts)}/{fmt(dur)}  falls:{events}{'  ⚠FALL' if is_fall else ''}",end="",flush=True)

    cap.release()
    if wr: wr.release()
    print()
    if apath and apath.exists(): print(f"  ✓  {apath.name}  ({apath.stat().st_size/1024**2:.1f}MB)")

    df=pd.DataFrame(rows)
    if WRITE_CLIPS and not df.empty:
        evs=segment_events(df)
        if evs:
            banner(f"Writing {len(evs)} clip(s)")
            write_clips(video_path,evs,out_dir/"fall_clips",fps,fw,fh,fall_model,dur,two_step_det,mode_lbl)
    return df

# ── Metrics ───────────────────────────────────────────────────────────────────
def compute_metrics(df, mode_lbl):
    if df.empty: return {}
    total=len(df); ff=int(df["is_fall"].sum())
    fdf=df[df["is_fall"]==1]
    evs=0
    if not fdf.empty:
        ts=sorted(fdf["timestamp_s"].tolist()); evs=1
        for a,b in zip(ts,ts[1:]):
            if b-a>EVENT_GAP: evs+=1
    mc=int(df["consecutive"].max()) if "consecutive" in df else 0
    ac=round(float(fdf["consecutive"].mean()) if not fdf.empty and "consecutive" in fdf else 0,2)
    return {"mode":mode_lbl,"total_frames":total,"fall_frames":ff,"clean_frames":total-ff,
            "fall_rate":round(ff/total,4) if total else 0,"distinct_events":evs,
            "mean_confidence":round(float(fdf["confidence"].mean()) if not fdf.empty else 0,4),
            "max_confidence": round(float(fdf["confidence"].max())  if not fdf.empty else 0,4),
            "class_counts":df["dominant_class"].value_counts().to_dict(),
            "max_consecutive":mc,"avg_consecutive":ac,
            "duration_s":round(float(df["timestamp_s"].max()),1)}

# ── Plots ─────────────────────────────────────────────────────────────────────
def plot_all(df, metrics, out_dir, stem, mode_lbl):
    time=df["timestamp_s"].values; conf=df["confidence"].values; fall=df["is_fall"].values.astype(bool)

    # Timeline
    fig,axes=plt.subplots(2,1,figsize=(15,7),facecolor="white",gridspec_kw={"height_ratios":[3,1],"hspace":0.10})
    ax=axes[0]; ax.set_facecolor(C_LIGHT)
    ax.fill_between(time,conf,where=fall, color=C_RED, alpha=0.30,label="Fall")
    ax.fill_between(time,conf,where=~fall,color=C_BLUE,alpha=0.10,label="No Fall")
    ax.plot(time,conf,color=C_BLUE,linewidth=0.9,alpha=0.6)
    ax.scatter(time[fall], conf[fall], color=C_RED,  s=20,zorder=3,alpha=0.9)
    ax.scatter(time[~fall],conf[~fall],color=C_GRAY, s=4, zorder=2,alpha=0.20)
    ax.axhline(CONF_THRESHOLD,color=C_AMBER,linestyle="--",linewidth=1.5,label=f"Threshold {CONF_THRESHOLD:.0%}")
    if metrics.get("mean_confidence",0)>0:
        ax.axhline(metrics["mean_confidence"],color=C_RED,linestyle=":",linewidth=1.0,label=f"Mean {metrics['mean_confidence']:.0%}")
    ax.set_ylabel("Confidence",fontsize=11); ax.set_ylim(-0.02,1.06); ax.set_xlim(time[0],time[-1])
    ax.set_title(f"{stem}  [{mode_lbl}]  —  {metrics['distinct_events']} events | {metrics['fall_rate']:.1%} fall rate",fontsize=12,weight="bold",pad=8)
    ax.legend(fontsize=9,loc="upper right"); ax.grid(axis="y",color="white"); ax.spines[["top","right"]].set_visible(False); ax.set_xticklabels([])
    ax2=axes[1]; ax2.set_facecolor(C_LIGHT)
    window=max(1,int(ROLL_WIN_SEC/(FRAME_SKIP/30.0)/FRAME_SKIP))
    roll=pd.Series(fall.astype(float)).rolling(window,min_periods=1).mean().values
    ax2.fill_between(time,roll,color=C_RED,alpha=0.35); ax2.plot(time,roll,color=C_RED,linewidth=1.0)
    ax2.set_ylabel(f"{ROLL_WIN_SEC}s\nRate",fontsize=9); ax2.set_xlabel("Timestamp",fontsize=11)
    ax2.set_ylim(0,1.05); ax2.set_xlim(time[0],time[-1])
    xticks=ax2.get_xticks(); ax2.set_xticks(xticks); ax2.set_xticklabels([fmt(t) for t in xticks],fontsize=8)
    ax2.spines[["top","right"]].set_visible(False); ax2.grid(axis="y",color="white")
    plt.tight_layout(); fig.savefig(str(out_dir/"timeline.png"),dpi=150,bbox_inches="tight"); plt.close(fig)

    # Confidence dist
    fig,axes=plt.subplots(1,2,figsize=(12,5),facecolor="white")
    ax=axes[0]; ax.set_facecolor(C_LIGHT)
    for cls,color in HEX_COLS.items():
        sub=df[df["dominant_class"]==cls]
        if not sub.empty: ax.hist(sub["confidence"],bins=20,alpha=0.65,color=color,label=cls,edgecolor="white")
    ax.axvline(CONF_THRESHOLD,color=C_AMBER,linestyle="--",linewidth=2)
    ax.set_xlabel("Confidence"); ax.set_ylabel("Frames"); ax.set_title(f"Confidence by Class  [{mode_lbl}]",fontsize=11,weight="bold")
    ax.legend(); ax.spines[["top","right"]].set_visible(False); ax.grid(axis="y",color="white")
    ax2=axes[1]; ax2.set_facecolor(C_LIGHT)
    fdf=df[df["is_fall"]==1]; nfdf=df[df["is_fall"]==0]
    if not fdf.empty:  ax2.hist(fdf["confidence"], bins=20,alpha=0.7,color=C_RED,  label=f"Fall (n={len(fdf)})", edgecolor="white")
    if not nfdf.empty: ax2.hist(nfdf["confidence"],bins=20,alpha=0.5,color=C_GREEN,label=f"No Fall (n={len(nfdf)})",edgecolor="white")
    ax2.axvline(CONF_THRESHOLD,color=C_AMBER,linestyle="--",linewidth=2)
    ax2.set_xlabel("Confidence"); ax2.set_title("Fall vs No-Fall",fontsize=11,weight="bold")
    ax2.legend(); ax2.spines[["top","right"]].set_visible(False); ax2.grid(axis="y",color="white")
    plt.tight_layout(); fig.savefig(str(out_dir/"confidence_dist.png"),dpi=150,bbox_inches="tight"); plt.close(fig)

    # Class breakdown
    fig,axes=plt.subplots(1,2,figsize=(12,5),facecolor="white")
    ax=axes[0]; ax.set_facecolor(C_LIGHT)
    cc=df["dominant_class"].value_counts()
    bars=ax.bar(cc.index,cc.values,color=[HEX_COLS.get(k,C_GRAY) for k in cc.index],edgecolor="white",width=0.55)
    for bar,val in zip(bars,cc.values):
        ax.text(bar.get_x()+bar.get_width()/2,bar.get_height()+0.5,str(val),ha="center",va="bottom",fontsize=12,weight="bold")
    ax.set_title(f"Class Distribution  [{mode_lbl}]",fontsize=12,weight="bold"); ax.set_ylabel("Frames")
    ax.spines[["top","right"]].set_visible(False); ax.grid(axis="y",color="white")
    ax2=axes[1]; ax2.set_facecolor(C_LIGHT)
    combo=df["classes_seen"].value_counts().head(10)
    ax2.barh(combo.index,combo.values,color=C_AMBER,edgecolor="white")
    for i,(idx,val) in enumerate(combo.items()): ax2.text(val+0.2,i,str(val),va="center",fontsize=10)
    ax2.set_xlabel("Frame Count"); ax2.set_title("Class Combinations",fontsize=12,weight="bold")
    ax2.spines[["top","right"]].set_visible(False); ax2.grid(axis="x",color="white"); ax2.invert_yaxis()
    plt.tight_layout(); fig.savefig(str(out_dir/"class_breakdown.png"),dpi=150,bbox_inches="tight"); plt.close(fig)

    # Dashboard
    fig=plt.figure(figsize=(15,9),facecolor="white")
    fig.suptitle(f"Fall Detection Summary — {stem}  [{mode_lbl}]",fontsize=14,weight="bold",y=0.98)
    gs=gridspec.GridSpec(2,4,figure=fig,hspace=0.45,wspace=0.35,top=0.91,bottom=0.08,left=0.06,right=0.97)
    cards=[("Fall Rate",f"{metrics['fall_rate']:.1%}",C_RED if metrics["fall_rate"]>0.1 else C_GREEN),
           ("Events",   str(metrics["distinct_events"]),C_RED),
           ("Mean Conf",f"{metrics['mean_confidence']:.1%}",C_GREEN if metrics["mean_confidence"]>=0.60 else C_AMBER),
           ("Duration", fmt(metrics["duration_s"]),C_GRAY)]
    for i,(label,value,color) in enumerate(cards):
        ax=fig.add_subplot(gs[0,i])
        ax.set_facecolor(color+"18"); ax.set_xlim(0,1); ax.set_ylim(0,1); ax.axis("off")
        ax.text(0.5,0.60,value,ha="center",va="center",fontsize=24,weight="bold",color=color)
        ax.text(0.5,0.18,label,ha="center",va="center",fontsize=9)
        for sp in ax.spines.values(): sp.set_color(color); sp.set_linewidth(2); sp.set_visible(True)
    ax_tl=fig.add_subplot(gs[1,:3]); ax_tl.set_facecolor(C_LIGHT)
    ax_tl.fill_between(time,conf,where=fall, color=C_RED, alpha=0.45,label="Fall")
    ax_tl.fill_between(time,conf,where=~fall,color=C_BLUE,alpha=0.12,label="No Fall")
    ax_tl.plot(time,conf,color=C_DARK,linewidth=0.7,alpha=0.45)
    ax_tl.axhline(CONF_THRESHOLD,color=C_AMBER,linestyle="--",linewidth=1.2,label=f"Thresh {CONF_THRESHOLD:.0%}")
    ax_tl.set_ylim(0,1.05); ax_tl.set_xlim(time[0],time[-1])
    ax_tl.set_xlabel("Time",fontsize=9); ax_tl.set_ylabel("Confidence",fontsize=9)
    ax_tl.set_title("Detection Timeline",fontsize=10,weight="bold")
    xticks=ax_tl.get_xticks(); ax_tl.set_xticks(xticks); ax_tl.set_xticklabels([fmt(t) for t in xticks],fontsize=7)
    ax_tl.legend(fontsize=8,loc="upper right"); ax_tl.spines[["top","right"]].set_visible(False); ax_tl.grid(axis="y",color="white")
    ax_bar=fig.add_subplot(gs[1,3]); ax_bar.set_facecolor(C_LIGHT)
    bars2=ax_bar.bar(cc.index,cc.values,color=[HEX_COLS.get(k,C_GRAY) for k in cc.index],edgecolor="white",width=0.55)
    for bar,val in zip(bars2,cc.values):
        ax_bar.text(bar.get_x()+bar.get_width()/2,bar.get_height()+max(cc.values)*0.01,str(val),ha="center",va="bottom",fontsize=9,weight="bold")
    ax_bar.set_title("Class Distribution",fontsize=10,weight="bold")
    ax_bar.spines[["top","right"]].set_visible(False); ax_bar.grid(axis="y",color="white")
    fig.text(0.5,0.01,f"Model: {FALL_MODEL_PATH}  |  Mode: {mode_lbl}  |  Threshold: {CONF_THRESHOLD:.0%}  |  Generated: {datetime.now().strftime('%d %b %Y %H:%M')}",
             ha="center",fontsize=8,color=C_GRAY)
    fig.savefig(str(out_dir/"summary_dashboard.png"),dpi=150,bbox_inches="tight"); plt.close(fig)
    print(f"  ✓  plots saved → {out_dir.name}/")

# ── Excel ─────────────────────────────────────────────────────────────────────
def export_excel(df, metrics, out_path, vname, mode_lbl):
    wb=Workbook(); ws1=wb.active; ws1.title="Detections"
    hf=PatternFill("solid",fgColor="2E5496"); hfont=Font(bold=True,color="FFFFFF",size=11)
    ctr=Alignment(horizontal="center",vertical="center")
    thin=Border(left=Side(style="thin",color="CCCCCC"),right=Side(style="thin",color="CCCCCC"),
                top=Side(style="thin",color="CCCCCC"),bottom=Side(style="thin",color="CCCCCC"))
    headers=[("Frame",9),("Timestamp",11),("Time(s)",10),("Fall?",10),("Alert",12),
             ("Dom.Class",14),("Confidence",13),("Consecutive",13),
             ("Classes",24),("#Boxes",8),("Fall",7),("Sit",7),("Walk",7)]
    for ci,(h,w) in enumerate(headers,1):
        cell=ws1.cell(row=1,column=ci,value=h)
        cell.fill=hf; cell.font=hfont; cell.alignment=ctr; cell.border=thin
        ws1.column_dimensions[get_column_letter(ci)].width=w
    FF=PatternFill("solid",fgColor="FFCDD2"); AF=PatternFill("solid",fgColor="F5F5F5"); BF=PatternFill("solid",fgColor="FFFFFF")
    for ri,row in enumerate(df.itertuples(index=False),2):
        isf=row.is_fall==1; rfill=FF if isf else (AF if ri%2==0 else BF)
        vals=[row.frame_number,row.timestamp_str,row.timestamp_s,"FALL" if isf else "Clear",
              getattr(row,"alert_level","—"),row.dominant_class,row.confidence,
              getattr(row,"consecutive",0),row.classes_seen,row.num_boxes,
              "✓" if row.has_fall else "","✓" if row.has_sit else "","✓" if row.has_walk else ""]
        for ci,val in enumerate(vals,1):
            cell=ws1.cell(row=ri,column=ci,value=val)
            cell.alignment=ctr; cell.border=thin; cell.fill=rfill
            cell.font=Font(size=10,bold=(isf and ci==4),color=("C0392B" if isf and ci==4 else "000000"))
    ws1.freeze_panes="A2"; ws1.auto_filter.ref=ws1.dimensions
    # Summary sheet
    ws2=wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width=28; ws2.column_dimensions["B"].width=22
    sf2=PatternFill("solid",fgColor="E8EEF7")
    for r,(label,value) in enumerate([
        ("VIDEO",""),("File",vname),("Mode",mode_lbl),("Duration",fmt(metrics["duration_s"])),
        ("",""),("RESULTS",""),
        ("Total frames",metrics["total_frames"]),("Fall frames",metrics["fall_frames"]),
        ("Fall rate",f"{metrics['fall_rate']:.2%}"),("Events",metrics["distinct_events"]),
        ("",""),("CONFIDENCE",""),
        ("Mean",f"{metrics['mean_confidence']:.4f}"),("Max",f"{metrics['max_confidence']:.4f}"),
        ("",""),("TEMPORAL",""),
        ("Max consecutive",metrics.get("max_consecutive",0)),
        ("Avg consecutive",metrics.get("avg_consecutive",0)),
        ("Persistence frames",FALL_PERSIST),
        ("",""),("CONFIG",""),
        ("Threshold",CONF_THRESHOLD),("Frame skip",FRAME_SKIP),
        ("Generated",datetime.now().strftime("%d %b %Y %H:%M")),
    ],1):
        ca=ws2.cell(row=r,column=1,value=label); cb=ws2.cell(row=r,column=2,value=value)
        thin2=Border(left=Side(style="thin",color="CCCCCC"),right=Side(style="thin",color="CCCCCC"),
                     top=Side(style="thin",color="CCCCCC"),bottom=Side(style="thin",color="CCCCCC"))
        ca.border=thin2; cb.border=thin2; cb.alignment=Alignment(horizontal="center")
        if label in ("VIDEO","RESULTS","CONFIDENCE","TEMPORAL","CONFIG"):
            ca.font=Font(bold=True,size=11,color="2E5496"); ca.fill=sf2; cb.fill=sf2
    wb.save(out_path); print(f"  ✓  {Path(out_path).name}")

# ── Comparison report ─────────────────────────────────────────────────────────
def write_comparison(all_results, out_path):
    lines=["="*80,"  COS30018 — Mode Comparison Report",
           f"  {datetime.now().strftime('%d %b %Y %H:%M')}","="*80,"",
           "  MODES:  1-step | 1-step+temporal | 2-step | 2-step+temporal","",
           "  KEY: Does 2-step / temporal reduce false positives while keeping real falls?","",
           f"  {'Video':<28} {'Mode':<22} {'FallRate':>9} {'Events':>7} {'MeanConf':>9} {'MaxConsec':>10} {'Delta':>8}",
           "  "+"─"*84]
    from itertools import groupby
    videos={}
    for vname,mode,m in all_results: videos.setdefault(vname,[]).append((mode,m))
    for vname,results in videos.items():
        lines.append(f"\n  ── {vname}")
        base=next((m for mode,m in results if mode=="1-step"),None)
        for mode,m in results:
            fr=m["fall_rate"]; evs=m["distinct_events"]; mc=m["mean_confidence"]; maxc=m.get("max_consecutive",0)
            delta=f"{fr-base['fall_rate']:+.1%}" if base and mode!="1-step" else "baseline"
            lines.append(f"  {vname:<28} {mode:<22} {fr:>8.1%} {evs:>7} {mc:>9.3f} {maxc:>10}  {delta}")
    lines+=["\n"+"="*80,"  INTERPRETATION",
            "  Fall rate lower in 2-step/temporal  → fewer false positives",
            "  Distinct events similar across modes → genuine falls preserved",
            "  Max consecutive higher               → sustained detection (less flickering)",
            "="*80]
    Path(out_path).write_text("\n".join(lines),encoding="utf-8")
    print("\n"+"\n".join(lines[-20:]))  # print tail

# ── Run one mode ──────────────────────────────────────────────────────────────
def run_mode(fall_model, person_model, vpath, vout, two_step_on, temporal_on):
    parts=["2-step" if two_step_on else "1-step"]
    if temporal_on: parts.append("temporal")
    mode_lbl=" + ".join(parts)
    out_dir=vout/mode_lbl.replace(" ","_").replace("+","plus")
    out_dir.mkdir(parents=True,exist_ok=True)
    two_step_det=TwoStep(person_model,fall_model) if (two_step_on and person_model) else None
    df=process_video(fall_model,vpath,out_dir,two_step_det,temporal_on,mode_lbl)
    if df.empty: return mode_lbl,{}
    m=compute_metrics(df,mode_lbl); m["video_name"]=vpath.name
    banner(f"Results [{mode_lbl}]")
    print(f"  Fall rate  : {m['fall_rate']:.1%}  ({m['fall_frames']} frames)"); print(f"  Events     : {m['distinct_events']}")
    print(f"  Mean conf  : {m['mean_confidence']:.3f}"); print(f"  Max consec : {m.get('max_consecutive',0)}")
    print(f"  Classes    : {m['class_counts']}")
    banner("Saving outputs")
    plot_all(df,m,out_dir,vpath.stem,mode_lbl)
    export_excel(df,m,str(out_dir/"detections.xlsx"),vpath.name,mode_lbl)
    return mode_lbl,m

# ── Entry point ───────────────────────────────────────────────────────────────
def main():
    parser=argparse.ArgumentParser(description="COS30018 — Fall Detection Video Tester")
    parser.add_argument("--two-step",    action="store_true")
    parser.add_argument("--temporal",    action="store_true")
    parser.add_argument("--compare-all", action="store_true",help="Run all 4 modes")
    args=parser.parse_args()

    print("\n"+"═"*62); print("  COS30018 — Comparative Fall Detection Video Testing"); print("═"*62)
    if not Path(FALL_MODEL_PATH).exists(): print(f"[ERROR] {FALL_MODEL_PATH}"); sys.exit(1)
    banner(f"Loading fall model: {FALL_MODEL_PATH}")
    fall_model=YOLO(FALL_MODEL_PATH); print(f"  Classes: {fall_model.names}")
    person_model=None
    if args.two_step or args.compare_all:
        if Path(PERSON_MODEL_PATH).exists():
            banner(f"Loading person model: {PERSON_MODEL_PATH}")
            person_model=YOLO(PERSON_MODEL_PATH); print(f"  Classes: {person_model.names}")
        else: print(f"  [WARN] Person model not found — 2-step modes skipped")

    videos=get_videos(VIDEOS_DIR); banner(f"Found {len(videos)} video(s)")
    for v in videos: print(f"  {v.name}")
    out_root=Path(RESULTS_DIR); out_root.mkdir(exist_ok=True)
    all_results=[]

    if args.compare_all:
        modes=[(False,False),(False,True),(True,False),(True,True)]
        for vpath in videos:
            vout=out_root/vpath.stem; vout.mkdir(exist_ok=True)
            for ts_on,temp_on in modes:
                if ts_on and not person_model: continue
                mode_lbl,m=run_mode(fall_model,person_model,vpath,vout,ts_on,temp_on)
                if m: all_results.append((vpath.name,mode_lbl,m))
        if all_results:
            write_comparison(all_results,str(out_root/"comparison_report.txt"))
    else:
        for vpath in videos:
            vout=out_root/vpath.stem; vout.mkdir(exist_ok=True)
            ts_on=args.two_step and person_model is not None
            mode_lbl,m=run_mode(fall_model,person_model,vpath,vout,ts_on,args.temporal)
            if m: all_results.append((vpath.name,mode_lbl,m))

    banner("All done"); print(f"  Results → {out_root.resolve()}")
    if args.compare_all: print(f"  Comparison → {out_root/'comparison_report.txt'}")

if __name__=="__main__": main()